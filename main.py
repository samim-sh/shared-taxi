from scipy.interpolate import interp1d
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import pandas
import collections
import itertools
import time
import re
import sys
import os
import psutil
import plotly
import openpyxl
from openpyxl.styles import Font, alignment, borders


# <editor-fold desc="Input number of taxis and riders">
No_Taxi = int(input('Enter #Taxi= '))
No_riders = int(input('Enter #Rider= '))
# </editor-fold>


# <editor-fold desc="generator">
def generator(beginning, ending):
    while beginning < ending:
        yield beginning
        beginning += 1
# </editor-fold>


# <editor-fold desc="print">
# Disable Prints
def block_print():
    sys.stdout = open(os.devnull, 'w')


# Restore Prints
def enable_print():
    sys.stdout = sys.__stdout__
# </editor-fold>


# <editor-fold desc="Return 2 random integers within a range(start, end)">
def random(start, end):
    first, second = np.random.randint(start, end), np.random.randint(start, end)
    return first, second
# </editor-fold>


# <editor-fold desc="Generate a random rider at location (x_h, y_h) in a random home area">
def sel_home_area():
    random_home_area = np.random.choice(Home_list)
    x_h, y_h = random(0, 45)
    x_min_home_area, x_max_home_area = Home[random_home_area][0][0], Home[random_home_area][0][1]
    y_min_home_area, y_max_home_area = Home[random_home_area][1][0], Home[random_home_area][1][1]
    # Check until (x_h, y_h) stand within selected random home area
    while x_min_home_area > x_h or x_h > x_max_home_area or y_min_home_area > y_h or y_h > y_max_home_area:
        x_h, y_h = random(0, 45)
    return x_h, y_h, random_home_area
# </editor-fold>


# <editor-fold desc="Generate a random rider at location (x_w, y_w) in a random work area">
def sel_work_area():
    random_work_area = np.random.choice(Work_list)
    x_w, y_w = random(0, 45)
    x_min_work_area, x_max_work_area = Work[random_work_area][0][0], Work[random_work_area][0][1]
    y_min_work_area, y_max_work_area = Work[random_work_area][1][0], Work[random_work_area][1][1]
    # Check until (x_w, y_w) stand within selected random work area
    while x_min_work_area > x_w or x_w > x_max_work_area or y_min_work_area > y_w or y_w > y_max_work_area:
        x_w, y_w = random(0, 45)
    return x_w, y_w, random_work_area
# </editor-fold>


api = 0


# <editor-fold desc="Calculate and return distance between 2 points (e.g. (x1, y1) and (x2, y2))">
def distance(node1, node2):
    global api
    api += 1
    return abs(node1[0]-node2[0])+abs(node1[1]-node2[1])
# </editor-fold>


# <editor-fold desc="Removing selected taxi and riders from taxi, origin and destination dictionaries">
def del_dict_item(selected_taxi, comb_dic, index_dic):
    del Taxis[selected_taxi]
    for each_rider in comb_dic[selected_taxi][index_dic[selected_taxi][0]]:
        del riders_O[each_rider]
        del riders_D[re.sub(r'o', 'd', each_rider)]
# </editor-fold>


# <editor-fold desc="Return traveled dist for each rider respectively, from pick up to drop off point in sharing mode">
def finding_distance_sharing(origin_list, destination_list):
    temp_lst = []  # temp_lst store traveled distance for each rider until drop off point
    sequence_pick_up_drop_off = origin_list + destination_list
    for rider in origin_list:
        rider_sequence = sequence_pick_up_drop_off[origin_list.index(rider):
                                                       sequence_pick_up_drop_off.index(re.sub(r'o', 'd', rider))+1]
        each_one_traveled_dist = 0
        for each in range(len(rider_sequence) - 1):
            if re.findall(r'[a-zA-Z]+', rider_sequence[each])[0] == 'pd':
                prior = riders_D[rider_sequence[each]]
            else:
                prior = riders_O[rider_sequence[each]]
            if re.findall(r'[a-zA-Z]+', rider_sequence[each + 1])[0] == 'pd':
                following = riders_D[rider_sequence[each + 1]]
            else:
                following = riders_O[rider_sequence[each + 1]]
            each_one_traveled_dist += distance(prior, following)
        temp_lst.append(each_one_traveled_dist)
    return temp_lst
# </editor-fold>


# <editor-fold desc="generate hex color">
def color():
    def hex():
        return np.random.randint(250)
    return '#{0:02X}{1:02X}{2:02X}'.format(hex(), hex(), hex())
# </editor-fold>


# <editor-fold desc="create areas, riders, taxis">
# ----------------------------------------------------------------------------------------------------------------------
#                                              Creating Home and Work Area
# ----------------------------------------------------------------------------------------------------------------------
Home = {}
Work = {}
area = 0
for y in range(0, 31, 15):
    y1, y2 = y, y+15
    for x in range(0, 31, 15):
        x1, x2 = x, x+15
        if area == 4 or area == 8:
            Work["w{0}".format(area+1)] = [(x1, x2), (y1, y2)]
        else:
            Home["h{0}".format(area+1)] = [(x1, x2), (y1, y2)]
        area += 1
Home_list = list(Home.keys())
Work_list = list(Work.keys())


# ----------------------------------------------------------------------------------------------------------------------
#                                                Generate Taxi Randomly
# ----------------------------------------------------------------------------------------------------------------------
Taxis = {}
for taxi in generator(1, No_Taxi+1):
    if taxi < (No_Taxi-No_Taxi/5):
        x, y, sel_home = sel_home_area()
        Taxis["t{0}".format(taxi)] = (x, y)
    else:
        x, y, sel_work = sel_work_area()
        Taxis["t{0}".format(taxi)] = (x, y)


# ----------------------------------------------------------------------------------------------------------------------
#                                                Generate Rider Randomly
# ----------------------------------------------------------------------------------------------------------------------
Riders_O = {}
riders_O = {}
riders_D = {}
person = 1
while len(riders_O) != No_riders:
    # Generate /--one fifth--/ of #riders work to home
    if person > int(No_riders * (1-1/5)):
        x, y, sel_work = sel_work_area()
        xx, yy, sel_home = sel_home_area()
        if distance((x, y), (xx, yy)) > 10:
            Riders_O["p{0}_".format(person) + sel_work + '_to_' + sel_home] = (x, y)  # Origin_to_Destination
            riders_O["po{0}".format(person)] = (x, y)  # Origin
            riders_D["pd{0}".format(person)] = (xx, yy)  # Destination
            person += 1

    # Generate /--one fifth--/ of (#riders*(4 / 5) home to home
    elif person > int(No_riders * (1-1/5)**2):
        x, y, sel_home = sel_home_area()
        xx, yy, sel_home_sec = sel_home_area()
        while sel_home_sec == sel_home:
            xx, yy, sel_home_sec = sel_home_area()
        if distance((x, y), (xx, yy)) > 10:
            Riders_O["p{0}_".format(person) + sel_home + '_to_' + sel_home_sec] = (x, y)  # Origin_to_Destination
            riders_O["po{0}".format(person)] = (x, y)  # Origin
            riders_D["pd{0}".format(person)] = (xx, yy)  # Destination
            person += 1

    # Generate /--#riders*(4 / 5)--/ of #riders home to work
    elif person > 0:
        x, y, sel_home = sel_home_area()
        xx, yy, sel_work = sel_work_area()
        if distance((x, y), (xx, yy)) > 10:
            Riders_O["p{0}_".format(person) + sel_home + '_to_' + sel_work] = (x, y)  # Origin_to_Destination
            riders_O["po{0}".format(person)] = (x, y)  # Origin
            riders_D["pd{0}".format(person)] = (xx, yy)  # Destination
            person += 1
# </editor-fold>

# <editor-fold desc="Distance based on individual mode(each rider travel with own vehicle)">
traveled_distance = {}
for each_person in riders_O:
    traveled_distance[each_person] = distance(riders_O[each_person], riders_D[re.sub(r'o', 'd', each_person)])
# </editor-fold>

# <editor-fold desc="plot body 1">
# ----------------------------------------------------------------------------------------------------------------------
#                                                        Plotting
# ----------------------------------------------------------------------------------------------------------------------
fig, ax = plt.subplots()
alpha = [1.9, .6, .25, .15]
patterns = ['+', 'x', '-']
for each_area in generator(1, 10):
    if each_area == 5 or each_area == 9:
        x, y, w, h = Work["w{0}".format(each_area)][0][0], \
                     Work["w{0}".format(each_area)][1][0], \
                     Work["w{0}".format(each_area)][0][1] - Work["w{0}".format(each_area)][0][0], \
                     Work["w{0}".format(each_area)][1][1] - Work["w{0}".format(each_area)][1][0]
        rect = patches.Rectangle((x, y), w, h, facecolor='red', alpha=alpha[3], )
        ax.add_patch(rect)
    else:
        x, y, w, h = Home["h{0}".format(each_area)][0][0], \
                     Home["h{0}".format(each_area)][1][0], \
                     Home["h{0}".format(each_area)][0][1]-Home["h{0}".format(each_area)][0][0], \
                     Home["h{0}".format(each_area)][1][1]-Home["h{0}".format(each_area)][1][0]
        rect = patches.Rectangle((x, y), w, h, alpha=alpha[3], )
        ax.add_patch(rect)


# 1 scatter plot with different text at each data point
i = 0
numbers = np.arange(1, No_Taxi+1)
for x, y in list(Taxis.values()):
    ax.text(x, y, numbers[i])
    i += 1


ax.axis([0, 60, 0, 45])
ax.minorticks_on()

ax.grid(b=True, which="minor", linestyle="--", color='r', alpha=0.4)
ax.grid(b=True, which="major", linestyle="-", linewidth=1, color="b", alpha=0.4)

ax.set_axisbelow(True)
ax.set_xticks(np.arange(0, 46, 15))
ax.set_yticks(np.arange(0, 46, 15))
ax.set_xticks(np.arange(0, 46, 1), minor=True)
ax.set_yticks(np.arange(0, 46, 1), minor=True)

blue_patch = patches.Patch(color='blue', label="Home's Area")
red_patch = patches.Patch(color='red', label="Work's Area")
red_dot = ax.scatter(*zip(*list(Riders_O.values())), marker='.', c='r', s=10, label="Rider's Origin")
blue_dot = ax.scatter(*zip(*list(riders_D.values())), marker='.', c='blue', s=10, label="Rider's Destination")
green_star = ax.scatter(*zip(*list(Taxis.values())), marker='*', c='fuchsia', s=40, label="Taxi's Location")

for xyo, xyd in zip(list(riders_O.values()), list(riders_D.values())):
    ax.annotate("", xytext=xyo, xy=xyd, arrowprops=dict(color=color(), arrowstyle="->"))
ax.legend(handles=[blue_patch, red_patch, red_dot, blue_dot, green_star], loc='best', title='Legend')
# </editor-fold>

# <editor-fold desc="define variables">
# Copy main dictionaries for using next for loops (SIF)
Taxis_copy, riders_O_copy, riders_D_copy = Taxis.copy(), riders_O.copy(), riders_D.copy()

# y_Traveled_Distance == Sharing Mode
# y_Traveled_distance == Sharing Mode + Individual Mode
# y_traveled_distance == Individual Mode
No_riders_list, No_zero_list, No_one_list, No_two_list, No_three_list, No_four_list = [], [], [], [], [], []
y_Traveled_Distance, y_Traveled_distance, y_traveled_distance = [], [], []
api_list, api_fix, elapsed_time, rss = [], api, [], []
# </editor-fold>

# <editor-fold desc="Create text file for writing print lines">
file = open('/home/samim/PycharmProjects/SharedTaxi/log.txt', 'w')
# Writing in log.txt file
file.write('{0}\n\n{1}\n\n{2}\n\n{3}\n\n'.format(Riders_O, riders_O_copy, riders_D_copy, Taxis_copy))
# </editor-fold>

# <editor-fold desc="Create excel file for results (output.xlsx)">
excel_writer = pandas.ExcelWriter('/home/samim/PycharmProjects/SharedTaxi/output.xlsx')
# </editor-fold>

# <editor-fold desc="summary of sif sheet in active_cars">
# Data frame for active cars in system
df_active_cars = pandas.DataFrame(columns=['#zero',
                                           '#one',
                                           '#two',
                                           '#three',
                                           '#four',
                                           'Sum_taxi',
                                           'Sum_rider',
                                           'Nan',
                                           'Missing',
                                           'Total'])
# </editor-fold>

Start = time.time()
# For loop to find optimum SIF ======================================================== Sharing importance factor == SIF
for SIF in [round(_, 1) for _ in np.linspace(1, 2.5, 16)]:
    # Writing in log.txt file
    file.write((('\n\n*\n*\n SIF = {0}\n*\n*\n'.format(SIF)).replace(' ', ' ' * 40, 1)).replace('*', '*' * 100, 4))

    # <editor-fold desc="sheet for each sif">
    # Data frame for results with headers
    df = pandas.DataFrame(columns=['Taxi',
                                   'Sequence at the origin', 'Sequence at the destination',
                                   'to first rider', 'origin', 'transfer', 'destination',
                                   '(sharing)', ' ', ' ', ' ',
                                   '(individual)', ' ', ' ', ' ',
                                   'Sharing/individual', ' ', ' ', ' '])
    df.loc[' '] = [' ',
                   ' ', ' ',
                   ' ', ' ', ' ', ' ',
                   'P1', 'P2', 'P3', 'P4',
                   'P1', 'P2', 'P3', 'P4',
                   'P1', 'P2', 'P3', 'P4']
    # </editor-fold>

    # <editor-fold desc="define variables">
    start_sif_loop = time.time()
    api = api_fix
    # For each SIF we should Retrieve Taxis, riders_O and riders_D dictionaries
    Taxis, riders_O, riders_D = Taxis_copy.copy(), riders_O_copy.copy(), riders_D_copy.copy()
    No_zero, No_one, No_two, No_three, No_four = 0, 0, 0, 0, 0  # Number of taxis with this riders
    Traveled_Distance = {}  # Distance each taxi travel in sharing mode
    nan_rider = []
    # </editor-fold>

    # Check all taxis
    for each_taxi in generator(0, len(Taxis)):
        check = True  # Check if there is at least one taxi in comb_locO_com_cons_min dictionary
        check_3 = True  # Check if there is at least one taxi in comb_locO_com_cons_min_3 dictionary
        # Writing in log.txt file
        # file.write((('\n\n*\n Taxi = {}\n'.format(each_taxi+1)).replace(' ', ' ' * 40, 1)).replace('*', '*' * 100, 2))

        # <editor-fold desc="if rider be in catchment area of taxi, it will be put in taxi's list">
        taxi_catchment = collections.defaultdict(list)
        for j in Taxis:
            catchment_area = patches.Circle(Taxis[j], radius=3)
            for each_person in riders_O:
                if catchment_area.contains_point(riders_O[each_person]):
                    taxi_catchment[j].append(each_person)
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('taxi_catchment', dict(taxi_catchment)))
        # </editor-fold>

        # <editor-fold desc="4-combination">
        # --------------------------------------------------------------------------------------------------------------
        #                                                   4
        # --------------------------------------------------------------------------------------------------------------
        # Find 4-combination of riders for each taxi
        comb = collections.defaultdict(list)
        for i in taxi_catchment:
            if len(taxi_catchment[i]) >= 4:
                for j in itertools.combinations(taxi_catchment[i], 4):
                    comb[i].append(j)
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb', dict(comb)))

        # Origin 4-coordinates (from rider_O dictionary) for 4-combination of riders for each taxi
        comb_locO = collections.defaultdict(list)
        for i in comb:
            for j in range(len(comb[i])):
                list_temporary = []
                for k in range(len(comb[i][j])):
                    list_temporary.append(riders_O[comb[i][j][k]])
                comb_locO[i].append(list_temporary)
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locO', dict(comb_locO)))

        # Destination 4-coordinates (from rider_D dictionary) for 4-combination of riders for each taxi
        comb_locD = collections.defaultdict(list)
        for i in comb:
            for j in range(len(comb[i])):
                list_temporary = []
                for k in range(len(comb[i][j])):
                    list_temporary.append(riders_D[re.sub(r'o', 'd', comb[i][j][k])])
                comb_locD[i].append(list_temporary)
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locD', dict(comb_locD)))

        # Calculating center of points for each group of riders for each taxi in origin
        comb_locO_com = collections.defaultdict(list)
        for i in comb_locO:
            comb_locO_com[i].append((np.array(comb_locO[i]).mean(axis=1)).tolist())
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locO_com', dict(comb_locO_com)))

        # Calculating center of points for each group of riders for each taxi in destination
        comb_locD_com = collections.defaultdict(list)
        for i in comb_locO:
            comb_locD_com[i].append((np.array(comb_locD[i]).mean(axis=1)).tolist())
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locD_com', dict(comb_locD_com)))

        # --------------------------------------------------------------------------------------------------------------
        #                   Calculating constraints (const_origin + const_taxi + const_destination)
        # --------------------------------------------------------------------------------------------------------------
        comb_locO_com_cons = collections.defaultdict(list)
        for i in comb_locO_com:
            for j in range(len(comb_locO[i])):
                Constraint = 0
                for k in range(len(comb_locO[i][j])):
                    Constraint += distance(comb_locO_com[i][0][j], comb_locO[i][j][k]) + \
                                  distance(comb_locD_com[i][0][j], comb_locD[i][j][k])
                comb_locO_com_cons[i].append(Constraint + distance(comb_locO_com[i][0][j], Taxis[i]))
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locO_com_cons', dict(comb_locO_com_cons)))

        # Minimum values of Constraints
        comb_locO_com_cons_min = collections.defaultdict(list)
        for i in comb_locO_com_cons:
            comb_locO_com_cons_min[i].append(min(comb_locO_com_cons[i]))
        # If comb_locO_com_cons dictionary be empty then comb_locO_com_cons_min will be empty
        # i.e. minimum calculating get ValueError so we can't calculate four in Comparision minimum of 3 and 4 riders
        try:
            minimum = min(comb_locO_com_cons_min, key=comb_locO_com_cons_min.get)
            # Writing in log.txt file
            # file.write('{0:41} ::= {1}\n'.format('comb_locO_com_cons_min', dict(comb_locO_com_cons_min)))
            file.write('{0:41} ::= {1}\n'.format('4', minimum))
        except ValueError:
            check = False

        # Index of minimum values of Constraints
        comb_locO_com_cons_min_index = collections.defaultdict(list)
        for i in comb_locO_com_cons:
            index = comb_locO_com_cons[i].index(min(comb_locO_com_cons[i]))
            comb_locO_com_cons_min_index[i].append(index)
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locO_com_cons_min_index', dict(comb_locO_com_cons_min_index)))
        # </editor-fold>

        # <editor-fold desc="3-combination">
        # --------------------------------------------------------------------------------------------------------------
        #                                                  3
        # --------------------------------------------------------------------------------------------------------------
        # Find 3-combination of riders for each taxi
        comb_3 = collections.defaultdict(list)
        for i in taxi_catchment:
            if len(taxi_catchment[i]) >= 3:
                for j in itertools.combinations(taxi_catchment[i], 3):
                    comb_3[i].append(j)
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_3', dict(comb_3)))

        # Origin 3-coordinates (from rider_O dictionary) for 3-combination of riders for each taxi
        comb_locO_3 = collections.defaultdict(list)
        for i in comb_3:
            for j in range(len(comb_3[i])):
                list_temporary = []
                for k in range(len(comb_3[i][j])):
                    list_temporary.append(riders_O[comb_3[i][j][k]])
                comb_locO_3[i].append(list_temporary)
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locO_3', dict(comb_locO_3)))

        # Destination 3-coordinates (from rider_D dictionary) for 3-combination of riders for each taxi
        comb_locD_3 = collections.defaultdict(list)
        for i in comb_3:
            for j in range(len(comb_3[i])):
                list_temporary = []
                for k in range(len(comb_3[i][j])):
                    list_temporary.append(riders_D[re.sub(r'o', 'd', comb_3[i][j][k])])
                comb_locD_3[i].append(list_temporary)
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locD_3', dict(comb_locD_3)))

        # Calculating center of points for each group of riders for each taxi in origin
        comb_locO_com_3 = collections.defaultdict(list)
        for i in comb_locO_3:
            comb_locO_com_3[i].append((np.array(comb_locO_3[i]).mean(axis=1)).tolist())
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locO_com_3', dict(comb_locO_com_3)))

        # Calculating center of points for each group of riders for each taxi in destination
        comb_locD_com_3 = collections.defaultdict(list)
        for i in comb_locO_3:
            comb_locD_com_3[i].append((np.array(comb_locD_3[i]).mean(axis=1)).tolist())
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locD_com_3', dict(comb_locD_com_3)))

        # --------------------------------------------------------------------------------------------------------------
        #                   Calculating constraints_3 (const_origin + const_taxi + const_destination)
        # --------------------------------------------------------------------------------------------------------------
        comb_locO_com_cons_3 = collections.defaultdict(list)
        for i in comb_locO_com_3:
            for j in range(len(comb_locO_3[i])):
                Constraint = 0
                for k in range(len(comb_locO_3[i][j])):
                    Constraint += distance(comb_locO_com_3[i][0][j], comb_locO_3[i][j][k]) + \
                                 distance(comb_locD_com_3[i][0][j], comb_locD_3[i][j][k])
                # ================================================================================================== SIF
                comb_locO_com_cons_3[i].append(SIF*(Constraint + distance(comb_locO_com_3[i][0][j], Taxis[i])))
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locO_com_cons_3', dict(comb_locO_com_cons_3)))

        # Minimum values of Constraints_3
        comb_locO_com_cons_min_3 = collections.defaultdict(list)
        for i in comb_locO_com_cons_3:
            comb_locO_com_cons_min_3[i].append(min(comb_locO_com_cons_3[i]))
        # If comb_locO_com_cons_3 dictionary be empty then comb_locO_com_cons_min_3 will be empty
        # i.e. minimum_3 calculating get ValueError so we can't calculate three in Comparision minimum of 3 and 4 riders
        try:
            minimum_3 = min(comb_locO_com_cons_min_3, key=comb_locO_com_cons_min_3.get)
            # Writing in log.txt file
            # file.write('{0:41} ::= {1}\n'.format('comb_locO_com_cons_min_3', dict(comb_locO_com_cons_min_3)))
            file.write('{0:41} ::= {1}\n'.format('3', minimum_3))
        except ValueError:
            check_3 = False

        # Index of minimum values of Constraints_3
        comb_locO_com_cons_min_index_3 = collections.defaultdict(list)
        for i in comb_locO_com_cons_3:
            index = comb_locO_com_cons_3[i].index(min(comb_locO_com_cons_3[i]))
            comb_locO_com_cons_min_index_3[i].append(index)
        # Writing in log.txt file
        # file.write('{0:41} ::= {1}\n'.format('comb_locO_com_cons_min_index_3', dict(comb_locO_com_cons_min_index_3)))
        # </editor-fold>

        # <editor-fold desc="terminate the SIF loop">
        # --------------------------------------------------------------------------------------------------------------
        # If comb_locO_com_cons and comb_locO_com_cons_3 dictionaries be empty then terminate loop
        # i.e. if there are no riders in catchment area of any taxis then stop this loop and go for new SIF
        if not bool(comb_locO_com_cons) and not bool(comb_locO_com_cons_3):
            # Writing in log.txt file
            file.write('\n{}\n\n'.format('There is no possible taxi to choose.'))
            break
        # </editor-fold>

        # <editor-fold desc="Comparision minimum of 3 and 4 riders">
        # --------------------------------------------------------------------------------------------------------------
        #                                  Comparision minimum of 3 and 4 riders
        # --------------------------------------------------------------------------------------------------------------
        try:
            four = comb_locO_com_cons[minimum][comb_locO_com_cons_min_index[minimum][0]] / 4
        except (IndexError, NameError):
            four = 1000
        try:
            three = comb_locO_com_cons_3[minimum_3][comb_locO_com_cons_min_index_3[minimum_3][0]] / 3
        except (IndexError, NameError):
            three = 1000
        # Writing in log.txt file
        file.write('{0:41} ::= {1}\n'.format('four', four))
        file.write('{0:41} ::= {1}\n'.format('three', three))
        # </editor-fold>

        # We want to know do we have any taxi with 4 riders, by boolean operator (check)
        if four <= three and check:
            # <editor-fold desc="if four choose">
            # Writing in log.txt file
            file.write('Taxi {0} is chosen with 4 rider.\n'.format(minimum))
            No_four += 1

            # ----------------------------------------------------------------------------------------------------------
            #                                       Shortest path (Start)
            # ----------------------------------------------------------------------------------------------------------
            # path_finding dictionary has two keys {1:[origin sequence], 2:[destination sequence]}
            path_finding = collections.defaultdict(list)
            # Find chosen combination from comb dictionary
            chosen_rider = comb[minimum][comb_locO_com_cons_min_index[minimum][0]]
            for each_key in chosen_rider:
                path_finding[1].append(each_key)
                path_finding[2].append(re.sub(r'o', 'd', each_key))
            # Writing in log.txt file
            file.write('{0:41} ::= {1}\n'.format('path_finding', path_finding))

            # Find permutation of riders for chosen taxi in both origin and destination
            path_finding_permutation = collections.defaultdict(list)
            for i in path_finding:
                for j in itertools.permutations(path_finding[i]):
                    path_finding_permutation[i].append(j)
            # Writing in log.txt file
            file.write('{0:41} ::= {1}\n'.format('path_finding_permutation', path_finding_permutation))

            # Calculating traveled distance for each permutation in ============================================= origin
            # path_finding_permutation_dist = {1:[]}
            path_finding_permutation_dist = collections.defaultdict(list)
            for j in range(len(path_finding_permutation[1])):
                sum_dist = 0
                for k in range(len(path_finding_permutation[1][j]) - 1):
                    sum_dist += distance(riders_O[path_finding_permutation[1][j][k]],
                                         riders_O[path_finding_permutation[1][j][k + 1]])
                path_finding_permutation_dist[1].append(sum_dist +
                                                        distance(Taxis[minimum],
                                                                 riders_O[path_finding_permutation[1][j][0]]))

            # ----------------------------------------------------------------------------------------------------------
            # ----------------------------------------------------------------------------------------------------------
            # Calculating traveled distance for each permutation in destination based on origin permutations
            # that are in path_finding_permutation[1] list, i.e. we try all origin sequence to get minimum distance
            # path_finding_permutation_dist = {1:[], 2:[[first_origin_min], [second_origin_min], [third_origin_min],..]}
            for i in range(len(path_finding_permutation[1])):
                list_temporary = []
                # Calculating traveled distance for each permutation in ==================================== destination
                for j in range(len(path_finding_permutation[2])):
                    sum_dist = 0
                    for k in range(len(path_finding_permutation[2][j]) - 1):
                        sum_dist += distance(riders_D[path_finding_permutation[2][j][k]],
                                             riders_D[path_finding_permutation[2][j][k + 1]])
                    list_temporary.append(sum_dist + distance(riders_O[path_finding_permutation[1][i][3]],
                                                              riders_D[path_finding_permutation[2][j][0]]))
                path_finding_permutation_dist[2].append(list_temporary)
            # Writing in log.txt file
            file.write('{0:41} ::= {1}\n'.format('path_finding_permutation_dist', path_finding_permutation_dist))

            # I don't care how many minimum are in each list of key(2):[  ] of path_finding_permutation_dist?
            # We have minimum in destination per origin's, min_2 = [number of permutation in origin]
            min_2 = [min(path_finding_permutation_dist[2][i]) for i in range(len(path_finding_permutation_dist[2]))]

            # By index_min_1 we can find which sequence of origin is the best for get the minimum traveled distance
            index_min_1 = min_2.index(min(min_2))
            path_min_1 = path_finding_permutation[1][index_min_1]
            min_1 = path_finding_permutation_dist[1][index_min_1]

            # Getting index of permutation that its traveled distance in (destination + origin) become the lowest and
            # Return rider's destination sequence in path_min_2 as a list
            index_min_2 = path_finding_permutation_dist[2][index_min_1].\
                index(min(path_finding_permutation_dist[2][index_min_1]))
            path_min_2 = path_finding_permutation[2][index_min_2]

            # ----------------------------------------------------------------------------------------------------------
            #                                       Put results in df (Data Frame)
            # ----------------------------------------------------------------------------------------------------------
            to_first_rider = distance(Taxis[minimum], riders_O_copy[path_min_1[0]])
            origin = min_1 - distance(Taxis[minimum], riders_O_copy[path_min_1[0]])
            transfer = distance(riders_O_copy[path_min_1[3]], riders_D_copy[path_min_2[0]])
            destination = min(min_2) - transfer

            p1_share, p2_share, p3_share, p4_share = finding_distance_sharing(path_min_1, path_min_2)

            p1_individual = distance(riders_O[path_min_1[0]], riders_D[re.sub(r'o', 'd', path_min_1[0])])
            p2_individual = distance(riders_O[path_min_1[1]], riders_D[re.sub(r'o', 'd', path_min_1[1])])
            p3_individual = distance(riders_O[path_min_1[2]], riders_D[re.sub(r'o', 'd', path_min_1[2])])
            p4_individual = distance(riders_O[path_min_1[3]], riders_D[re.sub(r'o', 'd', path_min_1[3])])

            # Share mode over individual mode for all chosen riders
            lst1 = np.array([p1_share, p2_share, p3_share, p4_share], dtype='float')
            lst2 = np.array([p1_individual, p2_individual, p3_individual, p4_individual], dtype='float')
            divide = lst1 / lst2
            divide[divide > 2] = np.NaN
            so1, so2, so3, so4 = divide

            # number_of_non_NaN_value = np.count_nonzero(~np.isnan(divide))
            number_of_NaN_value = np.count_nonzero(np.isnan(divide))

            # nan rider
            index_of_Nan = [i[0] for i in np.argwhere(np.isnan(divide)).tolist()]
            [nan_rider.append(path_min_1[i]) for i in index_of_Nan]
            index_of_non_Nan = [i[0] for i in np.argwhere(~np.isnan(divide)).tolist()]

            if number_of_NaN_value == 3 or number_of_NaN_value == 4:
                # Writing in log.txt file
                file.write((('\n-\n We had {} NaN values\n'.format(number_of_NaN_value)).
                            replace(' ', ' ' * 40, 1)).replace('-', '-' * 100, 2))
                No_four -= 1
                if number_of_NaN_value == 3:
                    No_one += 1
                else:
                    No_zero += 1
                pass
            elif number_of_NaN_value == 2 or number_of_NaN_value == 1:
                # Writing in log.txt file
                file.write((('\n-\n We had {} NaN values\n'.format(number_of_NaN_value)).
                            replace(' ', ' ' * 40, 1)).replace('-', '-' * 100, 2))
                No_four -= 1
                if number_of_NaN_value == 2:
                    No_two += 1
                else:
                    No_three += 1
                chosen_rider = [path_min_1[i] for i in index_of_non_Nan]
                # ------------------------------------------------------------------------------------------------------
                #                                       2_Shortest path (Start)
                # ------------------------------------------------------------------------------------------------------
                # path_finding dictionary has two keys {1:[origin sequence], 2:[destination sequence]}
                path_finding = collections.defaultdict(list)
                # Find chosen combination from comb dictionary
                for each_key in chosen_rider:
                    path_finding[1].append(each_key)
                    path_finding[2].append(re.sub(r'o', 'd', each_key))
                # Writing in log.txt file
                file.write('{0:41} ::= {1}\n'.format('path_finding', path_finding))

                # Find permutation of riders for chosen taxi in both origin and destination
                path_finding_permutation = collections.defaultdict(list)
                for i in path_finding:
                    for j in itertools.permutations(path_finding[i]):
                        path_finding_permutation[i].append(j)
                # Writing in log.txt file
                file.write('{0:41} ::= {1}\n'.format('path_finding_permutation', path_finding_permutation))

                # Calculating traveled distance for each permutation in ========================================= origin
                # path_finding_permutation_dist = {1:[]}
                path_finding_permutation_dist = collections.defaultdict(list)
                for j in range(len(path_finding_permutation[1])):
                    sum_dist = 0
                    for k in range(len(path_finding_permutation[1][j]) - 1):
                        sum_dist += distance(riders_O[path_finding_permutation[1][j][k]],
                                             riders_O[path_finding_permutation[1][j][k + 1]])
                    path_finding_permutation_dist[1].append(sum_dist +
                                                            distance(Taxis[minimum],
                                                                     riders_O[path_finding_permutation[1][j][0]]))
                # ------------------------------------------------------------------------------------------------------
                # ------------------------------------------------------------------------------------------------------
                # Calculating traveled distance for each permutation in destination based on origin permutations
                # that are in path_finding_permutation[1] list, i.e. we try all origin sequence to get minimum distance
                # path_finding_permutation_dist = {1:[],
                #                                  2:[[first_origin_min], [second_origin_min], [third_origin_min],..]}
                for i in range(len(path_finding_permutation[1])):
                    list_temporary = []
                    # Calculating traveled distance for each permutation in ================================ destination
                    for j in range(len(path_finding_permutation[2])):
                        sum_dist = 0
                        for k in range(len(path_finding_permutation[2][j]) - 1):
                            sum_dist += distance(riders_D[path_finding_permutation[2][j][k]],
                                                 riders_D[path_finding_permutation[2][j][k + 1]])
                        if number_of_NaN_value == 2:
                            list_temporary.append(sum_dist + distance(riders_O[path_finding_permutation[1][i][1]],
                                                                      riders_D[path_finding_permutation[2][j][0]]))
                        elif number_of_NaN_value == 1:
                            list_temporary.append(sum_dist + distance(riders_O[path_finding_permutation[1][i][2]],
                                                                      riders_D[path_finding_permutation[2][j][0]]))
                    path_finding_permutation_dist[2].append(list_temporary)
                # Writing in log.txt file
                file.write('{0:41} ::= {1}\n'.format('path_finding_permutation_dist', path_finding_permutation_dist))

                # I don't care how many minimum are in each list of key(2):[  ] of path_finding_permutation_dist?
                # We have minimum in destination per origin's, min_2 = [number of permutation in origin]
                min_2 = [min(path_finding_permutation_dist[2][i]) for i in range(len(path_finding_permutation_dist[2]))]

                # By index_min_1 we can find which sequence of origin is the best for get the minimum traveled distance
                index_min_1 = min_2.index(min(min_2))
                path_min_1 = path_finding_permutation[1][index_min_1]
                min_1 = path_finding_permutation_dist[1][index_min_1]

                # Getting index of permutation that its traveled distance in (destination + origin) become the lowest
                # and Return rider's destination sequence in path_min_2 as a list
                index_min_2 = path_finding_permutation_dist[2][index_min_1].\
                    index(min(path_finding_permutation_dist[2][index_min_1]))
                path_min_2 = path_finding_permutation[2][index_min_2]

                # ------------------------------------------------------------------------------------------------------
                #                                       Put results in df (Data Frame)
                # ------------------------------------------------------------------------------------------------------
                to_first_rider = distance(Taxis[minimum], riders_O_copy[path_min_1[0]])
                origin = min_1 - distance(Taxis[minimum], riders_O_copy[path_min_1[0]])
                if number_of_NaN_value == 2:
                    transfer = distance(riders_O_copy[path_min_1[1]], riders_D_copy[path_min_2[0]])

                    p1_share, p2_share = finding_distance_sharing(path_min_1, path_min_2)
                    p3_share = np.NaN
                    p4_share = np.NaN

                    p1_individual = distance(riders_O[path_min_1[0]],
                                             riders_D[re.sub(r'o', 'd', path_min_1[0])])
                    p2_individual = distance(riders_O[path_min_1[1]],
                                             riders_D[re.sub(r'o', 'd', path_min_1[1])])
                    p3_individual = np.NaN
                    p4_individual = np.NaN
                elif number_of_NaN_value == 1:
                    transfer = distance(riders_O_copy[path_min_1[2]], riders_D_copy[path_min_2[0]])

                    p1_share, p2_share, p3_share = finding_distance_sharing(path_min_1, path_min_2)
                    p4_share = np.NaN

                    p1_individual = distance(riders_O[path_min_1[0]],
                                             riders_D[re.sub(r'o', 'd', path_min_1[0])])
                    p2_individual = distance(riders_O[path_min_1[1]],
                                             riders_D[re.sub(r'o', 'd', path_min_1[1])])
                    p3_individual = distance(riders_O[path_min_1[2]],
                                             riders_D[re.sub(r'o', 'd', path_min_1[2])])
                    p4_individual = np.NaN
                destination = min(min_2) - transfer

                # Share mode over individual mode for all chosen riders
                lst1 = np.array([p1_share, p2_share, p3_share, p4_share], dtype='float')
                lst2 = np.array([p1_individual, p2_individual, p3_individual, p4_individual], dtype='float')
                divide = lst1 / lst2
                so1, so2, so3, so4 = divide

                # Compute distance in sharing mode(each taxi travel with riders)
                Traveled_Distance[minimum] = (min_1 + min(min_2))

            else:
                # Writing in log.txt file
                file.write((('\n-\n We had {} NaN values\n'.format(number_of_NaN_value)).
                            replace(' ', ' ' * 40, 1)).replace('-', '-' * 100, 2))
                # Compute distance in sharing mode(each taxi travel with riders)
                Traveled_Distance[minimum] = (min_1 + min(min_2))

            df.loc['{}'.format(each_taxi+1)] = [minimum,
                                                path_min_1, path_min_2,
                                                to_first_rider, origin, transfer, destination,
                                                p1_share, p2_share, p3_share, p4_share,
                                                p1_individual, p2_individual, p3_individual, p4_individual,
                                                round(so1, 2), round(so2, 2), round(so3, 2), round(so4, 2)]

            # Removing selected taxi and riders
            del_dict_item(minimum, comb, comb_locO_com_cons_min_index)

            # Writing in log.txt file
            file.write('{0:41} ::= {1}\n'.format('index_min_1', index_min_1))
            file.write('{0:41} ::= {1}\n'.format('path_min_1', path_min_1))
            file.write('{0:41} ::= {1}\n'.format('min_1', min_1))
            file.write('{0:41} ::= {1}\n'.format('index_min_2', index_min_2))
            file.write('{0:41} ::= {1}\n'.format('path_min_2', path_min_2))
            file.write('{0:41} ::= {1}\n'.format('min_2', min_2))
            # </editor-fold>

        # We want to know do we have any taxi with 3 riders, by boolean operator (check_3)
        elif four > three and check_3:
            # <editor-fold desc="if three choose">
            # Writing in log.txt file
            file.write('Taxi {0} is chosen with 3 rider.\n'.format(minimum_3))
            No_three += 1
            # ----------------------------------------------------------------------------------------------------------
            #                                      Shortest path_3 (Start)
            # ----------------------------------------------------------------------------------------------------------
            # path_finding dictionary has two keys {1:[origin sequence], 2:[destination sequence]}
            path_finding = collections.defaultdict(list)
            # Find chosen combination from comb_3 dictionary
            chosen_rider = comb_3[minimum_3][comb_locO_com_cons_min_index_3[minimum_3][0]]
            for each_key in chosen_rider:
                path_finding[1].append(each_key)
                path_finding[2].append(re.sub(r'o', 'd', each_key))
            # Writing in log.txt file
            file.write('{0:41} ::= {1}\n'.format('path_finding', path_finding))

            # Find permutation of riders for chosen taxi in both origin and destination
            path_finding_permutation = collections.defaultdict(list)
            for i in path_finding:
                for j in itertools.permutations(path_finding[i]):
                    path_finding_permutation[i].append(j)
            # Writing in log.txt file
            file.write('{0:41} ::= {1}\n'.format('path_finding_permutation', path_finding_permutation))

            # Calculating traveled distance for each permutation in ============================================= origin
            # path_finding_permutation_dist = {1:[]}
            path_finding_permutation_dist = collections.defaultdict(list)
            for j in range(len(path_finding_permutation[1])):
                sum_dist = 0
                for k in range(len(path_finding_permutation[1][j]) - 1):
                    sum_dist += distance(riders_O[path_finding_permutation[1][j][k]],
                                         riders_O[path_finding_permutation[1][j][k + 1]])
                path_finding_permutation_dist[1].append(sum_dist +
                                                        distance(Taxis[minimum_3],
                                                                 riders_O[path_finding_permutation[1][j][0]]))

            # ----------------------------------------------------------------------------------------------------------
            # ----------------------------------------------------------------------------------------------------------
            # Calculating traveled distance for each permutation in destination based on origin permutations
            # that are in path_finding_permutation[1] list, i.e. we try all origin sequence to get minimum distance
            # path_finding_permutation_dist = {1:[], 2:[[first_origin_min], [second_origin_min], [third_origin_min],..]}
            for i in range(len(path_finding_permutation[1])):
                list_temporary = []
                # Calculating traveled distance for each permutation in ==================================== destination
                for j in range(len(path_finding_permutation[2])):
                    sum_dist = 0
                    for k in range(len(path_finding_permutation[2][j]) - 1):
                        sum_dist += distance(riders_D[path_finding_permutation[2][j][k]],
                                             riders_D[path_finding_permutation[2][j][k + 1]])
                    list_temporary.append(sum_dist + distance(riders_O[path_finding_permutation[1][i][2]],
                                                              riders_D[path_finding_permutation[2][j][0]]))
                path_finding_permutation_dist[2].append(list_temporary)
            # Writing in log.txt file
            file.write('{0:41} ::= {1}\n'.format('path_finding_permutation_dist', path_finding_permutation_dist))

            # I don't care how many minimum are in each list of key(2):[  ] of path_finding_permutation_dist?
            # We have minimum in destination per origin's, min_2 = [number of permutation in origin]
            min_2 = [min(path_finding_permutation_dist[2][i]) for i in range(len(path_finding_permutation_dist[2]))]

            # By index_min_1 we can find which sequence of origin is the best for get the minimum traveled distance
            index_min_1 = min_2.index(min(min_2))
            path_min_1 = path_finding_permutation[1][index_min_1]
            min_1 = path_finding_permutation_dist[1][index_min_1]

            # Getting index of permutation that its traveled distance in (destination + origin) become the lowest and
            # Return rider's destination sequence in path_min_2 as a list
            index_min_2 = path_finding_permutation_dist[2][index_min_1].\
                index(min(path_finding_permutation_dist[2][index_min_1]))
            path_min_2 = path_finding_permutation[2][index_min_2]

            # ----------------------------------------------------------------------------------------------------------
            #                                       Put results in df (Data Frame)
            # ----------------------------------------------------------------------------------------------------------
            to_first_rider = distance(Taxis[minimum_3], riders_O_copy[path_min_1[0]])
            origin = min_1 - distance(Taxis[minimum_3], riders_O_copy[path_min_1[0]])
            transfer = distance(riders_O_copy[path_min_1[2]], riders_D_copy[path_min_2[0]])
            destination = min(min_2) - transfer

            p1_share, p2_share, p3_share = finding_distance_sharing(path_min_1, path_min_2)
            p4_share = np.NAN

            p1_individual = distance(riders_O[path_min_1[0]], riders_D[re.sub(r'o', 'd', path_min_1[0])])
            p2_individual = distance(riders_O[path_min_1[1]], riders_D[re.sub(r'o', 'd', path_min_1[1])])
            p3_individual = distance(riders_O[path_min_1[2]], riders_D[re.sub(r'o', 'd', path_min_1[2])])
            p4_individual = np.NAN

            # Share mode over individual mode for all chosen riders
            lst1 = np.array([p1_share, p2_share, p3_share], dtype='float')
            lst2 = np.array([p1_individual, p2_individual, p3_individual], dtype='float')
            divide = lst1 / lst2
            divide[divide > 2] = np.NaN
            so1, so2, so3 = divide
            so4 = np.NaN

            # number_of_non_NaN_value = np.count_nonzero(~np.isnan(divide))
            number_of_NaN_value = np.count_nonzero(np.isnan(divide))

            # nan rider
            index_of_Nan = [i[0] for i in np.argwhere(np.isnan(divide)).tolist()]
            [nan_rider.append(path_min_1[i]) for i in index_of_Nan]
            index_of_non_Nan = [i[0] for i in np.argwhere(~np.isnan(divide)).tolist()]

            if number_of_NaN_value == 2 or number_of_NaN_value == 3:
                # Writing in log.txt file
                file.write((('\n-\n We had {} NaN values\n'.format(number_of_NaN_value)).
                            replace(' ', ' ' * 40, 1)).replace('-', '-' * 100, 2))
                No_three -= 1
                if number_of_NaN_value == 2:
                    No_one += 1
                else:
                    No_zero += 1
                pass
            elif number_of_NaN_value == 1:
                # Writing in log.txt file
                file.write((('\n-\n We had {} NaN values\n'.format(number_of_NaN_value)).
                            replace(' ', ' ' * 40, 1)).replace('-', '-' * 100, 2))
                No_three -= 1
                No_two += 1
                chosen_rider = [path_min_1[i] for i in index_of_non_Nan]
                # ------------------------------------------------------------------------------------------------------
                #                                       2_Shortest path (Start)
                # ------------------------------------------------------------------------------------------------------
                # path_finding dictionary has two keys {1:[origin sequence], 2:[destination sequence]}
                path_finding = collections.defaultdict(list)
                # Find chosen combination from comb dictionary
                for each_key in chosen_rider:
                    path_finding[1].append(each_key)
                    path_finding[2].append(re.sub(r'o', 'd', each_key))
                # Writing in log.txt file
                file.write('{0:41} ::= {1}\n'.format('path_finding', path_finding))

                # Find permutation of riders for chosen taxi in both origin and destination
                path_finding_permutation = collections.defaultdict(list)
                for i in path_finding:
                    for j in itertools.permutations(path_finding[i]):
                        path_finding_permutation[i].append(j)
                # Writing in log.txt file
                file.write('{0:41} ::= {1}\n'.format('path_finding_permutation', path_finding_permutation))

                # Calculating traveled distance for each permutation in ========================================= origin
                # path_finding_permutation_dist = {1:[]}
                path_finding_permutation_dist = collections.defaultdict(list)
                for j in range(len(path_finding_permutation[1])):
                    sum_dist = 0
                    for k in range(len(path_finding_permutation[1][j]) - 1):
                        sum_dist += distance(riders_O[path_finding_permutation[1][j][k]],
                                             riders_O[path_finding_permutation[1][j][k + 1]])
                    path_finding_permutation_dist[1].append(sum_dist +
                                                            distance(Taxis[minimum_3],
                                                                     riders_O[path_finding_permutation[1][j][0]]))

                # ------------------------------------------------------------------------------------------------------
                # ------------------------------------------------------------------------------------------------------
                # Calculating traveled distance for each permutation in destination based on origin permutations
                # that are in path_finding_permutation[1] list, i.e. we try all origin sequence to get minimum distance
                # path_finding_permutation_dist = {1:[],
                #                                  2:[[first_origin_min], [second_origin_min], [third_origin_min],..]}
                for i in range(len(path_finding_permutation[1])):
                    list_temporary = []
                    # Calculating traveled distance for each permutation in ================================ destination
                    for j in range(len(path_finding_permutation[2])):
                        sum_dist = 0
                        for k in range(len(path_finding_permutation[2][j]) - 1):
                            sum_dist += distance(riders_D[path_finding_permutation[2][j][k]],
                                                 riders_D[path_finding_permutation[2][j][k + 1]])
                        list_temporary.append(sum_dist + distance(riders_O[path_finding_permutation[1][i][1]],
                                                                  riders_D[path_finding_permutation[2][j][0]]))
                    path_finding_permutation_dist[2].append(list_temporary)
                # Writing in log.txt file
                file.write('{0:41} ::= {1}\n'.format('path_finding_permutation_dist', path_finding_permutation_dist))

                # I don't care how many minimum are in each list of key(2):[  ] of path_finding_permutation_dist?
                # We have minimum in destination per origin's, min_2 = [number of permutation in origin]
                min_2 = [min(path_finding_permutation_dist[2][i]) for i in range(len(path_finding_permutation_dist[2]))]

                # By index_min_1 we can find which sequence of origin is the best for get the minimum traveled distance
                index_min_1 = min_2.index(min(min_2))
                path_min_1 = path_finding_permutation[1][index_min_1]
                min_1 = path_finding_permutation_dist[1][index_min_1]

                # Getting index of permutation that its traveled distance in (destination + origin) become the lowest
                # and Return rider's destination sequence in path_min_2 as a list
                index_min_2 = path_finding_permutation_dist[2][index_min_1].\
                    index(min(path_finding_permutation_dist[2][index_min_1]))
                path_min_2 = path_finding_permutation[2][index_min_2]

                # ------------------------------------------------------------------------------------------------------
                #                                       Put results in df (Data Frame)
                # ------------------------------------------------------------------------------------------------------
                to_first_rider = distance(Taxis[minimum_3], riders_O_copy[path_min_1[0]])
                origin = min_1 - distance(Taxis[minimum_3], riders_O_copy[path_min_1[0]])
                transfer = distance(riders_O_copy[path_min_1[1]], riders_D_copy[path_min_2[0]])
                destination = min(min_2) - transfer

                p1_share, p2_share = finding_distance_sharing(path_min_1, path_min_2)
                p3_share, p4_share = np.NAN, np.NaN

                p1_individual = distance(riders_O[path_min_1[0]], riders_D[re.sub(r'o', 'd', path_min_1[0])])
                p2_individual = distance(riders_O[path_min_1[1]], riders_D[re.sub(r'o', 'd', path_min_1[1])])
                p3_individual = np.NaN
                p4_individual = np.NAN

                # Share mode over individual mode for all chosen riders
                lst1 = np.array([p1_share, p2_share, p3_share, p4_share], dtype='float')
                lst2 = np.array([p1_individual, p2_individual, p3_individual, p4_individual], dtype='float')
                divide = lst1 / lst2
                so1, so2, so3, so4 = divide

                # Compute distance in sharing mode(each taxi travel with riders)
                Traveled_Distance[minimum_3] = (min_1 + min(min_2))

            else:
                # Writing in log.txt file
                file.write((('\n-\n We had {} NaN values\n'.format(number_of_NaN_value)).
                            replace(' ', ' ' * 40, 1)).replace('-', '-' * 100, 2))
                # Compute distance in sharing mode(each taxi travel with riders)
                Traveled_Distance[minimum_3] = (min_1 + min(min_2))

            df.loc['{}'.format(each_taxi + 1)] = [minimum_3,
                                                  path_min_1, path_min_2,
                                                  to_first_rider, origin, transfer, destination,
                                                  p1_share, p2_share, p3_share, p4_share,
                                                  p1_individual, p2_individual, p3_individual, p4_individual,
                                                  round(so1, 2), round(so2, 2), round(so3, 2), round(so4, 2)]

            # Removing selected taxi and riders
            del_dict_item(minimum_3, comb_3, comb_locO_com_cons_min_index_3)

            # Writing in log.txt file
            file.write('{0:41} ::= {1}\n'.format('index_min_1', index_min_1))
            file.write('{0:41} ::= {1}\n'.format('path_min_1', path_min_1))
            file.write('{0:41} ::= {1}\n'.format('min_1', min_1))
            file.write('{0:41} ::= {1}\n'.format('index_min_2', index_min_2))
            file.write('{0:41} ::= {1}\n'.format('path_min_2', path_min_2))
            file.write('{0:41} ::= {1}\n'.format('min_2', min_2))
            # </editor-fold>

    # <editor-fold desc="Sum up traveled distance of missing riders by sharing mode, so we have Traveled_distance">
    missing_rider = 0
    for i in riders_O:
        missing_rider += distance(riders_O[i], riders_D[re.sub(r'o', 'd', i)])
    # </editor-fold>

    # <editor-fold desc="Sum up traveled distance of riders who exceed (share to individual) ratio">
    nan_pass = 0
    for i in nan_rider:
        nan_pass += distance(riders_O_copy[i], riders_D_copy[re.sub(r'o', 'd', i)])
    # </editor-fold>

    # Writing in log.txt file
    file.write('{0:41} ::= {1}\n\n'.format('nan_rider {}'.format(SIF), nan_rider))
    file.write('{0:41} ::= {1}\n\n'.format('Traveled_Distance {}'.format(SIF), Traveled_Distance))
    file.write('{0:41} ::= {1}\n'.format('1-traveled_distance_individual{}'.format(SIF), sum(traveled_distance.values())))
    file.write('{0:41} ::= {1}\n'.format('2-Traveled_Distance_sharing {}'.format(SIF), sum(Traveled_Distance.values())))
    file.write('{0:41} ::= {1}\n'.format('3-missing_rider {}'.format(SIF), missing_rider))
    file.write('{0:41} ::= {1}\n'.format('4-nan_rider {}'.format(SIF), nan_pass))
    file.write('{0:41} ::= {1}\n'.format('5-Traveled_distance {} = (2+3+4)'.format(SIF),
                                         sum(Traveled_Distance.values())+missing_rider+nan_pass))

    # <editor-fold desc="save each sif as new sheet">
    df.to_excel(excel_writer, sheet_name='{}'.format(SIF))
    # </editor-fold>

    # <editor-fold desc="write each sif as a row in active_cars sheet">
    No_riders_list.append(No_two * 2 + No_three * 3 + No_four * 4)
    No_zero_list.append(No_zero)
    No_one_list.append(No_one)
    No_two_list.append(No_two)
    No_three_list.append(No_three)
    No_four_list.append(No_four)
    y_Traveled_Distance.append(sum(Traveled_Distance.values()))
    y_Traveled_distance.append(sum(Traveled_Distance.values()) + missing_rider)
    y_traveled_distance.append(sum(traveled_distance.values()))

    df_active_cars.loc['{}'.format(SIF)] = [No_zero,
                                            No_one,
                                            No_two,
                                            No_three,
                                            No_four,
                                            No_two+No_three+No_four,
                                            No_two*2+No_three*3+No_four*4,
                                            len(nan_rider),
                                            len(riders_O),
                                            No_two*2+No_three*3+No_four*4+len(nan_rider)+len(riders_O)]
    # </editor-fold>

    # <editor-fold desc="rss, api_list">
    rss.append(psutil.Process(os.getpid()).memory_info()[0])
    api_list.append(api)
    end_sif_loop = time.time()
    elapsed_time.append(round((end_sif_loop-start_sif_loop)/60, 2))
    # </editor-fold>
End = time.time()

# <editor-fold desc="save df_active_cars as active_cars sheet">
df_active_cars.to_excel(excel_writer, sheet_name='active_cars')
# </editor-fold>

# <editor-fold desc="save outpu.xlsx file">
excel_writer.save()
excel_writer.close()
# </editor-fold>

# <editor-fold desc="outpu.xlsx style !">
wb = openpyxl.load_workbook('/home/samim/PycharmProjects/SharedTaxi/output.xlsx')
for ws in ['{}'.format(round(i, 1)) for i in np.linspace(1, 2.5, 16)]:
    wb[ws].merge_cells('A1:A2')
    wb[ws].merge_cells('B1:B2')
    wb[ws].merge_cells('C1:C2')
    wb[ws].merge_cells('D1:D2')
    wb[ws].merge_cells('E1:E2')
    wb[ws].merge_cells('F1:F2')
    wb[ws].merge_cells('G1:G2')
    wb[ws].merge_cells('H1:H2')
    wb[ws].merge_cells('I1:L1')
    wb[ws].merge_cells('M1:P1')
    wb[ws].merge_cells('Q1:T1')
    bullshit = 0
    for row in ['{0}:{0}'.format(i) for i in range(1, 160)]:
        shit = 0
        for cell in wb[ws][row]:
            if bullshit in {0, 1}:
                cell.font = Font(bold=False, color='FF800080')
            if shit in {2, 3}:
                pass
            else:
                cell.alignment = alignment.Alignment(horizontal='center', vertical='top')
            if shit in {0, 4, 5, 6, 8, 9, 10, 12, 13, 14, 16, 17, 18}:
                pass
            else:
                cell.border = borders.Border(right=borders.Side(style='double'))
            shit += 1
        bullshit += 1
wb.save('/home/samim/PycharmProjects/SharedTaxi/output.xlsx')
wb.close()
# </editor-fold>

print('#Taxi: {0}, #Rider: {1}, solved in: {2} min'.format(No_Taxi, No_riders, round((End-Start)/60, 2)))
ax.set_title('#Taxi: {0}, #Rider: {1}, solved in: {2} min'.
             format(No_Taxi, No_riders, round((End-Start)/60, 2)))

# <editor-fold desc="plotly body">
# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++Number of active car-plotly
x = np.linspace(1, 2.5, 16)
# plot SIF versus #zero
trace0 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(No_zero_list),
    mode='lines+markers',
    name='SIF versus #zero'
)
# plot SIF versus #one
trace1 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(No_one_list),
    mode='lines+markers',
    name='SIF versus #one'
)
# plot SIF versus #two
trace2 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(No_two_list),
    mode='lines+markers',
    name='SIF versus #two'
)
# plot SIF versus #three
trace3 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(No_three_list),
    mode='lines+markers',
    name='SIF versus #three'
)
# plot SIF versus #four
trace4 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(No_four_list),
    mode='lines+markers',
    name='SIF versus #four'
)
# plot SIF versus (#two+#three+#four)
trace5 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(No_two_list) + np.array(No_three_list) + np.array(No_four_list),
    mode='lines+markers',
    name='SIF versus #two+#three+#four'
)
data = [trace0, trace1, trace2, trace3, trace4, trace5]
layout = plotly.graph_objs.Layout(
    title='<b>SIF versus taxis with 0, 1, 2, 3, 4 rider</b>',
    titlefont=dict(family='Comic Sans MS', size=18, color='#3c3c3c'),
    hovermode='closest',
    legend=dict(x=1, y=.5, traceorder='normal', font=dict(family='Courier', size=15, color='#000'),
                bgcolor='#FFFFFF', bordercolor='#E2E2E2', borderwidth=2),
    xaxis=dict(title='SIF', zeroline=True, gridwidth=2),
    yaxis=dict(title='Numbers of active Taxi', zeroline=True, gridwidth=2),
    showlegend=True
)
fig1 = plotly.graph_objs.Figure(data=data, layout=layout)
plotly.offline.plot(fig1, filename='active_taxi.html', auto_open=False)


# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++#Riders-plotly
# plot SIF versus #riders
trace0 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(No_riders_list),
    mode='lines+markers',
    name='SIF versus #riders'
)
data = [trace0]
layout = plotly.graph_objs.Layout(
    title='<b>SIF versus #riders</b>',
    titlefont=dict(family='Comic Sans MS', size=18, color='#3c3c3c'),
    hovermode='closest',
    legend=dict(x=1, y=.5, traceorder='normal', font=dict(family='Courier', size=15, color='#000'),
                bgcolor='#FFFFFF', bordercolor='#E2E2E2', borderwidth=2),
    xaxis=dict(title='SIF', zeroline=True, gridwidth=2),
    yaxis=dict(title='Numbers of riders who used SharedTaxi', zeroline=True, gridwidth=2),
    showlegend=True
)
fig1 = plotly.graph_objs.Figure(data=data, layout=layout)
plotly.offline.plot(fig1, filename='shared_riders.html', auto_open=False)


# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++traveled distance-plotly
# plot SIF versus Traveled_Distance
trace0 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(y_Traveled_Distance),
    mode='lines+markers',
    name='SIF versus Sharing Mode'
)
# plot SIF versus (Traveled_Distance+traveled_distance)
trace1 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(y_Traveled_distance),
    mode='lines+markers',
    name='SIF versus Sharing Mode+Individual Mode'
)
# plot SIF versus y_traveled_distance
trace2 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(y_traveled_distance),
    mode='lines+markers',
    name='SIF versus Individual Mode'
)
data = [trace0, trace1, trace2, trace3]
layout = plotly.graph_objs.Layout(
    title='<b>SIF versus traveled distance</b>',
    titlefont=dict(family='Comic Sans MS', size=18, color='#3c3c3c'),
    hovermode='closest',
    legend=dict(x=1, y=.5, traceorder='normal', font=dict(family='Courier', size=15, color='#000'),
                bgcolor='#FFFFFF', bordercolor='#E2E2E2', borderwidth=2),
    xaxis=dict(title='SIF', zeroline=True, gridwidth=2),
    yaxis=dict(title='Total traveled distance', zeroline=True, gridwidth=2),
    showlegend=True
)
fig2 = plotly.graph_objs.Figure(data=data, layout=layout)
plotly.offline.plot(fig2, filename='traveled_distance.html', auto_open=False)


# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++riders-plotly
# plot riders in origin
xo, yo = zip(*list(riders_O_copy.values()))
zo = []
for i in generator(0, len(list(riders_O_copy.values()))):
    Sum = sum([1 if j == list(riders_O_copy.values())[i] else 0 for j in list(riders_O_copy.values())[:i]])
    zo.append(Sum)
trace0 = plotly.graph_objs.Scatter3d(
    x=xo,
    y=yo,
    z=zo,
    text=list(riders_O_copy.keys()),
    textfont=dict(size=8),
    mode='markers+text',
    marker=dict(color='#00bfff', size=3, symbol='circle', line=dict(color='#0000ff', width=1), opacity=0.9),
    textposition='top center',
    name='Riders in Origin'
)

# plot riders in destination
xd, yd = map(list, zip(*list(riders_D_copy.values())))
zd = []
for i in generator(0, len(list(riders_D_copy.values()))):
    Sum = sum([1 if j == list(riders_D_copy.values())[i] else 0 for j in list(riders_D_copy.values())[:i]])
    zd.append(Sum)
trace1 = plotly.graph_objs.Scatter3d(
    x=xd,
    y=yd,
    z=zd,
    text=list(riders_D_copy.keys()),
    textfont=dict(size=8),
    mode='markers+text',
    marker=dict(color='#faaa0a', size=3, symbol='circle', line=dict(color='#9400d3', width=1), opacity=0.8),
    name='Riders in Destination'
)

# plot taxis
xt, yt = map(list, zip(*list(Taxis_copy.values())))
zt = []
for i in generator(0, len(list(Taxis_copy.values()))):
    Sum = sum([1 if j == list(Taxis_copy.values())[i] else 0 for j in list(Taxis_copy.values())[:i]])
    zt.append(Sum)
trace2 = plotly.graph_objs.Scatter3d(
    x=xt,
    y=yt,
    z=zt,
    text=list(Taxis_copy.keys()),
    textfont=dict(size=8),
    mode='markers+text',
    marker=dict(color='#800080', size=3, symbol='x', line=dict(color='#fa00fa', width=1), opacity=0.8),
    name='Taxi'
)
data = [trace0, trace1, trace2]
layout = plotly.graph_objs.Layout(
    title='Taxi: {0}, riders: {1}, solved in: {2} min'.format(No_Taxi, No_riders, round((End-Start)/60, 2)),
    titlefont=dict(family='Comic Sans MS', size=8, color='rgb(60, 60, 60)'),
    margin=dict(r=0, b=0, l=0, t=0),
    legend=dict(x=1, y=.5, traceorder='normal', font=dict(family='Courier', size=15, color='#000'),
                bgcolor='#FFFFFF', bordercolor='#E2E2E2', borderwidth=2),
    scene=dict(xaxis=dict(tickfont=dict(size=12), backgroundcolor="#c8c8e6", gridcolor="#ffffff", showbackground=True,
                          zerolinecolor="#ffffff"),
               yaxis=dict(tickfont=dict(size=12), backgroundcolor="#e6c8e6", gridcolor="#ffffff", showbackground=True,
                          zerolinecolor="#ffffff"),
               zaxis=dict(tickfont=dict(size=12), backgroundcolor="#e6e6c8", gridcolor="#ffffff", showbackground=True,
                          zerolinecolor="#ffffff"),
               dragmode="turntable",
               annotations=[dict(showarrow=True, z=3, text='<b>Taxi: {0}, riders: {1}, solved in: {2} min</b>'.
                                 format(No_Taxi, No_riders, round((End-Start)/60, 2)), textangle=0,
                                 font=dict(family='Comic Sans MS', size=18, color='#3c3c3c'),
                                 arrowcolor="rgba(255, 255, 255, .1)", ax=0, ay=-100)]  # rgba(255,255,255,0.1)=#fafafa
               )
)
fig3 = plotly.graph_objs.Figure(data=data, layout=layout)
plotly.offline.plot(fig3, filename='location.html', auto_open=False)


# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++api times-plotly
x = np.linspace(1, 2.5, 16)
# plot SIF versus api
trace0 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(api_list),
    mode='lines+markers',
    name='SIF versus api'
)
data = [trace0]
layout = plotly.graph_objs.Layout(
    title='<b>SIF versus api</b>',
    titlefont=dict(family='Comic Sans MS', size=18, color='#3c3c3c'),
    hovermode='closest',
    legend=dict(x=1, y=.5, traceorder='normal', font=dict(family='Courier', size=15, color='#000'),
                bgcolor='#FFFFFF', bordercolor='#E2E2E2', borderwidth=2),
    xaxis=dict(title='SIF', zeroline=True, gridwidth=2),
    yaxis=dict(title='Numbers of usage distance function', zeroline=True, gridwidth=2),
    showlegend=True
)
fig1 = plotly.graph_objs.Figure(data=data, layout=layout)
plotly.offline.plot(fig1, filename='api.html', auto_open=False)


# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++elapsed time-plotly
x = np.linspace(1, 2.5, 16)
# plot SIF versus Execution time
trace0 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(elapsed_time),
    mode='lines+markers',
    name='SIF versus Execution time'
)
data = [trace0]
layout = plotly.graph_objs.Layout(
    title='<b>SIF versus Execution time</b>',
    titlefont=dict(family='Comic Sans MS', size=18, color='#3c3c3c'),
    hovermode='closest',
    legend=dict(x=1, y=.5, traceorder='normal', font=dict(family='Courier', size=15, color='#000'),
                bgcolor='#FFFFFF', bordercolor='#E2E2E2', borderwidth=2),
    xaxis=dict(title='SIF', zeroline=True, gridwidth=2),
    yaxis=dict(title='Execution time', zeroline=True, gridwidth=2),
    showlegend=True
)
fig1 = plotly.graph_objs.Figure(data=data, layout=layout)
plotly.offline.plot(fig1, filename='elapsed_time.html', auto_open=False)


# ++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++rss-plotly
x = np.linspace(1, 2.5, 16)
# plot SIF versus rss
trace0 = plotly.graph_objs.Scatter(
    x=x,
    y=np.array(rss),
    mode='lines+markers',
    name='SIF versus rss'
)
data = [trace0]
layout = plotly.graph_objs.Layout(
    title='<b>SIF versus rss</b>',
    titlefont=dict(family='Comic Sans MS', size=18, color='#3c3c3c'),
    hovermode='closest',
    legend=dict(x=1, y=.5, traceorder='normal', font=dict(family='Courier', size=15, color='#000'),
                bgcolor='#FFFFFF', bordercolor='#E2E2E2', borderwidth=2),
    xaxis=dict(title='SIF', zeroline=True, gridwidth=2),
    yaxis=dict(title='rss', zeroline=True, gridwidth=2),
    showlegend=True
)
fig1 = plotly.graph_objs.Figure(data=data, layout=layout)
plotly.offline.plot(fig1, filename='rss.html', auto_open=False)


html_graphs = open("DASHBOARD.html", 'w')
html_graphs.write("<html><head></head><body>"+"\n")
html_graphs.write("<object data='location.html' width='900' height='700'></object>"+"\n")
html_graphs.write("<object data='active_taxi.html' width='900' height='700'></object>"+"\n")
html_graphs.write("<object data='shared_riders.html' width='900' height='700'></object>"+"\n")
html_graphs.write("<object data='traveled_distance.html' width='1100' height='800'></object>"+"\n")
html_graphs.write("<object data='api.html' width='900' height='700'></object>"+"\n")
html_graphs.write("<object data='elapsed_time.html' width='900' height='700'></object>"+"\n")
html_graphs.write("<object data='rss.html' width='900' height='700'></object>"+"\n")
html_graphs.write("</body></html>")
html_graphs.close()
# </editor-fold>


# <editor-fold desc="plot body">
# ==================================================================================================================plot
fig1, axes1 = plt.subplots(nrows=3, ncols=3)
x_new = np.linspace(x.min(), x.max(), 300)

# plot SIF versus #three
y = np.array(No_three_list)
f = interp1d(x, y, kind='quadratic')
y_smooth = f(x_new)
axes1[0, 0].plot(x_new, y_smooth)
axes1[0, 0].scatter(x, y)
axes1[0, 0].set_title('SIF versus #three')
axes1[0, 0].set_xlabel('SIF', fontsize=10)
axes1[0, 0].set_ylabel('Numbers of active Taxi', fontsize=10)

# plot SIF versus #four
y = np.array(No_four_list)
f = interp1d(x, y, kind='quadratic')
y_smooth = f(x_new)
axes1[0, 1].plot(x_new, y_smooth)
axes1[0, 1].scatter(x, y)
axes1[0, 1].set_title('SIF versus #four')

# plot SIF versus (#two+#three+#four)
temp_list = np.array(No_two_list) + np.array(No_three_list) + np.array(No_four_list)
y = temp_list
f = interp1d(x, y, kind='quadratic')
y_smooth = f(x_new)
axes1[0, 2].plot(x_new, y_smooth)
axes1[0, 2].scatter(x, y)
axes1[0, 2].set_title('SIF versus #two+#three+#four')

# plot SIF versus Traveled_Distance
y = np.array(y_Traveled_Distance)
f = interp1d(x, y, kind='quadratic')
y_smooth = f(x_new)
axes1[1, 0].plot(x_new, y_smooth)
axes1[1, 0].scatter(x, y)
axes1[1, 0].set_title('SIF versus Sharing Mode')
axes1[1, 0].set_xlabel('SIF', fontsize=10)
axes1[1, 0].set_ylabel('Total traveled distance', fontsize=10)

# plot SIF versus (Traveled_Distance+traveled_distance)
y = np.array(y_Traveled_distance)
f = interp1d(x, y, kind='quadratic')
y_smooth = f(x_new)
axes1[1, 1].plot(x_new, y_smooth)
axes1[1, 1].scatter(x, y)
axes1[1, 1].set_title('SIF versus Sharing Mode+Individual Mode')

# plot SIF versus y_traveled_distance
y = np.array(y_traveled_distance)
f = interp1d(x, y, kind='quadratic')
y_smooth = f(x_new)
axes1[1, 2].plot(x_new, y_smooth)
axes1[1, 2].scatter(x, y)
axes1[1, 2].set_title('SIF versus y_traveled_distance')

# plot SIF versus #Riders
y = np.array(No_riders_list)
f = interp1d(x, y, kind='quadratic')
y_smooth = f(x_new)
axes1[2, 0].plot(x_new, y_smooth)
axes1[2, 0].scatter(x, y)
axes1[2, 0].set_title('SIF versus #Riders')
axes1[2, 0].set_xlabel('SIF', fontsize=10)
axes1[2, 0].set_ylabel('Total #Riders', fontsize=10)

# plot SIF versus elapsed_time
y = np.array(elapsed_time)
f = interp1d(x, y, kind='quadratic')
y_smooth = f(x_new)
axes1[2, 1].plot(x_new, y_smooth)
axes1[2, 1].scatter(x, y)
axes1[2, 1].set_title('SIF versus elapsed_time')
axes1[2, 1].set_ylabel('Total elapsed_time', fontsize=10)

# plot SIF versus rss
y = np.array(rss)
f = interp1d(x, y, kind='quadratic')
y_smooth = f(x_new)
axes1[2, 2].plot(x_new, y_smooth)
axes1[2, 2].scatter(x, y)
axes1[2, 2].set_title('SIF versus rss')
axes1[2, 2].set_ylabel('Total rss', fontsize=10)

fig1.subplots_adjust(top=0.9, bottom=0.1, left=0.125, right=0.9, hspace=0.5, wspace=0.3)

# Writing in log.txt file
file.write('{0:16} ::= {1}\n'.format('virtual_memory()', psutil.virtual_memory()))
file.write('{0:16} ::= {1}\n'.format('swap_memory()', psutil.swap_memory()))
file.write('{0:16} ::= {1}\n'.format('memory_info()', psutil.Process(os.getpid()).memory_info()))
file.close()

plt.show()
fig.savefig('/home/samim/PycharmProjects/SharedTaxi/locations.png')
fig1.savefig('/home/samim/PycharmProjects/SharedTaxi/plot.png')
# </editor-fold>